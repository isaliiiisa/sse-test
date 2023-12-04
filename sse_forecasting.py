import pandas as pd
import numpy as np
import pickle

import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.figure_factory as ff
import plotly.express as px

import textwrap

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from sse_ml_utils import *

def select_values_subset(sample_df, value_col, team_col=None):
    level_columns = ['l1', 'l2', 'l3', 'l4', 
                     'l1_weight', 'l2_weight', 'l3_weight', 'l4_weight', 
                     'direction', 'value_range']
    if team_col:
        level_columns.append(team_col)
    return sample_df[level_columns + [value_col]]


def read_sample_data_targets(data):
    ### TO DO: add data checks
    sample_df = pd.read_excel(data)
    sample_df.columns = [str(c).strip().lower().replace(" ", "_") for c in sample_df.columns]
    
    capabilities_df = sample_df[sample_df.group == 'Capabilities']
    outcomes_df = sample_df[sample_df.group == 'Outcomes']
    return capabilities_df, outcomes_df



def prepare_outcomes_d(outcomes_df):
    outcomes_df['l4'] = outcomes_df['l4'].apply(lambda x: str(x).strip().lower().replace(" ", "_"))
    outcomes_d = dict(zip(outcomes_df.l4, outcomes_df.value))
    return outcomes_d



def normalize_weights(df, n_levels):
    df['l0'] = 'total_score'
    df['w_multiplied'] = 1
    for i in range(1, n_levels+1):
        w_keys   = df.drop_duplicates(subset=[f'l{i}'])[f'l{i}'].values

        w_values = df.drop_duplicates(subset=[f'l{i}'])[f'l{i}_weight'] / \
        df.drop_duplicates(subset=[f'l{i}']).\
        groupby([f'l{i-1}'])[f'l{i}_weight'].transform('sum') 

        w_dict = dict(zip(w_keys, w_values))

        df[f'l{i}_w'] = df[f'l{i}'].map(w_dict)

        df['w_multiplied'] = df['w_multiplied'] * df[f'l{i}_w']    
    return df


def customwrap(s,width=30):
    return "<br>".join(textwrap.wrap(s,width=width))
    

def assign_color(value):
    """
    Assigns a color based on the score value.

    :param value: The score value (expected to be a float or convertible to float).
    :return: Hex color code as a string.
    """
    try:
        score = float(value)
        if score > 0.75:
            return '63BF77'  # Green
        elif score > 0.5:
            return 'FFD458'  # Yellow
        elif score > 0.25:
            return 'FFB12A'  # Orange
        else:
            return 'FF5B59'  # Red
    except ValueError:
        return None 
    
    
    
def get_level_scores(df, score_col, levels_mapping_d, weights_mapping_d):
    agg_dfs_d = {}
    agg_dfs_d['l4'] = df
    agg_dfs_d['l4']['l4_weight'] = agg_dfs_d['l4']['l4'].map(weights_mapping_d['l4_weight'])
    agg_dfs_d['l4']['l4_weight'] = agg_dfs_d['l4'].groupby('l3')['l4_weight'].transform(lambda x: x/x.sum())
    
    agg_dfs_d['l4']['l4_weighted_score'] = agg_dfs_d['l4']['l4_weight'] * agg_dfs_d['l4']['l4_score']
    
    
    for i in range(3,0,-1):
        agg_dfs_d[f'l{i}'] = agg_dfs_d[f'l{i+1}'].groupby(f'l{i}').agg({f'l{i+1}_weighted_score':'sum'}).reset_index().\
        rename(columns={f'l{i+1}_weighted_score':f'l{i}_score'})
        
        ### Save level color mapping
        #color_mapping.update(dict(zip(agg_dfs_d[f'l{i}'][f'l{i}'], agg_dfs_d[f'l{i}'][f'l{i}_score'].apply(assign_color))))
                
        agg_dfs_d[f'l{i}'][f'l{i-1}'] = agg_dfs_d[f'l{i}'][f'l{i}'].map(levels_mapping_d[f'l{i}_l{i-1}_mapping'])  
        agg_dfs_d[f'l{i}'][f'l{i}_weight'] = agg_dfs_d[f'l{i}'][f'l{i}'].map(weights_mapping_d[f'l{i}_weight'])  
        agg_dfs_d[f'l{i}'][f'l{i}_weight'] = agg_dfs_d[f'l{i}'].groupby(f'l{i-1}')[f'l{i}_weight'].transform(lambda x: x/x.sum())
        agg_dfs_d[f'l{i}'][f'l{i}_weighted_score'] = agg_dfs_d[f'l{i}'][f'l{i}_weight'] * agg_dfs_d[f'l{i}'][f'l{i}_score']
    return agg_dfs_d



def create_color_mapping(agg_dfs_d):
    color_mapping = {}
    for i in range(1,5):
        color_mapping.update(dict(zip(agg_dfs_d[f'l{i}'][f'l{i}'], agg_dfs_d[f'l{i}'][f'l{i}_score'].apply(assign_color))))
    color_mapping['total_score'] = ''
    return color_mapping
        
    

def load_model(outcome_metric):
    model_filename =  f'temp_sse_{outcome_metric.replace(" ", "_")}_model_trained.sav'
    trained_model = pickle.load(open(model_filename, 'rb'))
    return trained_model



def make_sunburst_chart(sample_df, color_mapping):
    ### Fix hover template
    ### Add legend
    

    ### Adding text wrapping 
    for c in ['l0', 'l1', 'l2', 'l3', 'l4']:
        sample_df[c] = sample_df[c].map(customwrap)
    color_mapping = {customwrap(m):v for m,v in color_mapping.items()}
    
    ### Creating chart
    fig = px.sunburst(sample_df, path=['l0', 'l1', 'l2', 'l3', 'l4'],
                      values='w_multiplied',
                      )
    
    fig.update_traces(marker_colors=[color_mapping[cat] for cat in fig.data[-1].labels])
    fig.update_layout(margin=dict(l=0, r=0, t=20, b=0))
    fig.update_layout(
            autosize=True,
            width=500,
            height=500,
        )

    fig.update_xaxes(type='category', color="white")
    fig.update_layout(autosize=False, width=958, height=600)
    return fig



def prepare_output_table(df, value, levels_mapping_d, weights_mapping_d):
    df_subset = select_values_subset(df, value)
    df_subset_scores = calculate_L4_scores(df_subset, value)
    level_scores_d   = get_level_scores(df_subset_scores, 'l4_score', levels_mapping_d, weights_mapping_d)

    
    for i in range(1,4):
        score_d = dict(zip(level_scores_d[f'l{i}'][f'l{i}'], level_scores_d[f'l{i}'][f'l{i}_score']))
        df_subset_scores[f'l{i}_score'] = df_subset_scores[f'l{i}'].map(score_d)
        
    output_table = df_subset_scores[['l4', value, 'l4_score', 'l3', 'l3_score', 'l2', 'l2_score', 'l1', 'l1_score']]
    return output_table



def format_and_save_excel(df, file_path, value_col='value', n_levels=4):
    """
    Formats an Excel file with color based on scores, merges cells with the same values in level columns
    and their corresponding score columns, and aligns level values in the middle.

    :param df: Pandas DataFrame to be formatted and saved.
    :param n_levels: Number of levels in the DataFrame.
    :param file_path: Path to save the Excel file.
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Transfer the DataFrame to the Excel worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Identify and apply color formatting for score columns
            column_name = df.columns[c_idx - 1]  # Adjusting index for zero-based column index
            if '_score' in column_name:
                color = assign_color(value)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.number_format = '0.00'
            
            if value_col in column_name:
                cell.number_format = '0'
                
            # Apply middle alignment for level columns
            if any(column_name in [f'l{i}_score', value_col] for i in range(1, n_levels + 1)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if any(column_name in [f'l{i}'] for i in range(1, n_levels + 1)):
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Merging cells with same values in level columns and their corresponding score columns
    for level in range(1, n_levels + 1):
        level_column_name = f'l{level}'
        score_column_name = f'l{level}_score'

        if level_column_name in df.columns and score_column_name in df.columns:
            level_column_index = df.columns.get_loc(level_column_name) + 1  # Converting to 1-based index
            score_column_index = df.columns.get_loc(score_column_name) + 1

            level_column_letter = get_column_letter(level_column_index)
            score_column_letter = get_column_letter(score_column_index)

            previous_value = None
            start_row = None

            for row in range(1, ws.max_row + 1):
                cell_value = ws[level_column_letter + str(row)].value
                if cell_value != previous_value:
                    if start_row and row - start_row > 1:
                        ws.merge_cells(f'{level_column_letter}{start_row}:{level_column_letter}{row-1}')
                        ws.merge_cells(f'{score_column_letter}{start_row}:{score_column_letter}{row-1}')
                    start_row = row
                    previous_value = cell_value

            # Check for merge at the end of the column
            if start_row and row - start_row > 1:
                ws.merge_cells(f'{level_column_letter}{start_row}:{level_column_letter}{row}')
                ws.merge_cells(f'{score_column_letter}{start_row}:{score_column_letter}{row}')

    # Save the workbook to the file
    wb.save(file_path)
    
    
    
def plot_pipeline(df, value, levels_mapping_d, weights_mapping_d):
    df_subset = select_values_subset(df, value)
    df_subset_scores = calculate_L4_scores(df_subset, value)
    df_subset_scores = normalize_weights(df_subset_scores, 4)
    level_scores_d   = get_level_scores(df_subset_scores, 'l4_score', levels_mapping_d, weights_mapping_d)
    
    color_mapping  = create_color_mapping(level_scores_d)
    
    fig =  make_sunburst_chart(df_subset_scores, color_mapping)
    return fig
    
    
    
def plot_pipeline(df, value, levels_mapping_d, weights_mapping_d):
    df_subset = select_values_subset(df, value)
    df_subset_scores = calculate_L4_scores(df_subset, value)
    df_subset_scores = normalize_weights(df_subset_scores, 4)
    level_scores_d   = get_level_scores(df_subset_scores, 'l4_score', levels_mapping_d, weights_mapping_d)
    
    color_mapping  = create_color_mapping(level_scores_d)
    
    fig =  make_sunburst_chart(df_subset_scores, color_mapping)
    return fig



def forecast_uplift(outcome_metric, real_values, df_current, df_forecast):
    ### Model predictions
    uplift_mapping = {'speed':1.1, 'team_health':1.05, 'quality':1.21, 'efficiency':1.15}
    trained_model = load_model(outcome_metric)

    current_model_predicted = trained_model.predict(df_current)[0]
    forecast_model_predicted = trained_model.predict(df_forecast)[0]*uplift_mapping[outcome_metric]
    model_uplift = 100 * (forecast_model_predicted/ current_model_predicted - 1)
    
    ### Actual data
    current_outcome = real_values[outcome_metric]
    forecasted_outcome = current_outcome * (1 + model_uplift/100)
    
    ### model_error =  ((current_model_predicted - current_outcome) / current_model_predicted) * 100
    return model_uplift, current_outcome, forecasted_outcome #, model_error



def find_red_orange_areas(subset, i, n_metrics):
    #level_results = level_scores_d[f'l{i}']
    level_results = subset.copy()
    red_areas = level_results[level_results[f'l{i}_score']<=0.25].head(n_metrics)

    orange_n = n_metrics - red_areas.shape[0]
    if orange_n > 0:
        orange_areas = level_results[(level_results[f'l{i}_score']>0.25) & (level_results[f'l{i}_score']<=0.5)].sort_values(by=f'l{i}_score').head(orange_n)

    lagging_metrics = {}
    try: 
        lagging_metrics.update({m:'red' for m in red_areas[f'l{i}'].unique()})
    except:
        pass
    try: 
        lagging_metrics.update({m:'orange' for m in orange_areas[f'l{i}'].unique()})
    except:
        pass
    return lagging_metrics



def find_lagging_areas(df, value, levels_mapping_d, weights_mapping_d):
    df_subset = select_values_subset(df, value)
    df_subset_scores = calculate_L4_scores(df_subset, value)
    level_scores_d   = get_level_scores(df_subset_scores, 'l4_score', levels_mapping_d, weights_mapping_d)
    
    l3_lagging_d = find_red_orange_areas(level_scores_d['l3'], 3, n_metrics=5)

    l4_lagging_d = {}

    for lagging_l3 in l3_lagging_d.keys():
        subset = level_scores_d['l4'][level_scores_d['l4']['l3'] == lagging_l3]
        l4_lagging_d[lagging_l3] = find_red_orange_areas(subset, 4, n_metrics=3)
        
    return l3_lagging_d, l4_lagging_d



def plot_forecasted_metrics(data):
    # Colors for the bars and text
    color_current = '#001d2d'
    color_forecast = '#00abfa'
    color_metric_names = '#001F29'  # Color for metric names

    # Lists to store data for plotting
    labels = []
    current_values = []
    forecasted_values = []
    uplifts = []

    # Extract and process data from the dictionary
    for key, values in data.items():
        labels.append(key.capitalize())
        current_values.append(values['current outcome'])
        forecasted_values.append(values['forecasted outcome'])
        uplifts.append(f"{values['% uplift']:.2f}%")

    # Create traces for the current and forecasted values
    trace_current = go.Bar(x=labels, y=current_values, name='Current', marker_color=color_current)
    trace_forecast = go.Bar(x=labels, y=forecasted_values, name='Forecasted', marker_color=color_forecast)

    # Create the figure and add traces
    fig = go.Figure(data=[trace_current, trace_forecast])

    # Add annotations for uplifts
    for i, label in enumerate(labels):
        fig.add_annotation(
            x=label,
            y=forecasted_values[i],
            text=uplifts[i],
            showarrow=True,
            font=dict(
                size=14,
                color='black',
                family="Arial, bold"
            ),
            arrowhead=2,
            arrowsize=1,
            arrowwidth=2,
            arrowcolor='black',
            ax=20,  # Horizontal offset
            ay=-30  # Vertical offset
        )

    # Update layout for transparent background and bold font
    fig.update_layout(
        title='Forecasted Uplift',
        xaxis_title='',
        yaxis_title='Values',
        barmode='group',
        legend=dict(x=0.1, y=1.1, orientation='h'),
        plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
    )

    # Make metric names (x-axis labels) bold, size 20, color #001F29
    fig.update_xaxes(
        tickfont=dict(family="Arial, bold", size=20, color=color_metric_names),
        showgrid=False, gridwidth=1, gridcolor='lightgray'
    )
    fig.update_yaxes(showgrid=False, gridwidth=1, gridcolor='lightgray')

    return fig


def create_imp_plot(metric, data):
    # Extracting the specific metric data
    feature_importances = data[['feature_name', metric+'_importance_score', metric+'_importance_grade']]
    df = feature_importances .copy()
    df.columns = ['Feature', 'Importance', 'Grade']
    # Creating a DataFrame from the dictionary
    #df = pd.DataFrame(list(feature_importances.items()), columns=['Feature Id', 'Importance'])

    # Sorting values and selecting the top 20 features for better visualization
    df = df.sort_values('Importance', ascending=False)#.head(10)
    

    # Creating the bar chart
    fig = px.bar(df, x='Importance', y='Feature', orientation='h',
                 #title=f'Top 20 Feature Importances for {metric}',
                 color='Grade', color_discrete_map={
                                'A': '#001e2e',
                                'B': '#00abfb',
                                'C': '#0251ff',
                                'D': '#9ae7f2'
                                })

    # Adjusting layout for better readability
    fig.update_layout(yaxis={'categoryorder':'total ascending'},
                      xaxis_title="Importance Score",
                      yaxis_title="",
                      #showlegend=False,
                     paper_bgcolor='rgba(0,0,0,0)',
                     plot_bgcolor='rgba(0,0,0,0)')
    fig.update_layout(yaxis={"dtick":1},margin={"t":0,"b":0},height=360, width=1000)
    return fig
